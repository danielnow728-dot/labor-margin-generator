import re
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


# -----------------------------
# EMBEDDED INTERNAL CATEGORY MAP
# (extracted from your Internal catagory.xlsx)
# Keys must match inventory "section header" text (case-insensitive)
# Values must be one of: Labor, Rent, Material, Sub, Delivery, Other
# -----------------------------
INTERNAL_CATEGORY_MAP = {
    # From updated screenshot
    "RETENTION WITHHELD": "Other",
    "LABOR - YARD": "Labor",
    "LABOR - GENERAL MAINTENANCE": "Labor",
    "LABOR - INSULATION": "Labor",
    "LABOR - ASBESTOS": "Labor",
    "CONTROLLED INSURANCE DISCOUNT": "Other",
    "LABOR - SCAFFOLD": "Labor",
    "DEDUCT": "Other",
    "SWING STAGE - LABOR": "Labor",
    "SWING STAGE - RENTAL": "Rent",
    "SWING STAGE - DELIVERY/PICK-UP": "Delivery",
    "SWING STAGE - ENG. / DRAWINGS": "Sub",
    "SWING STAGE - PERM.EQUIP.SALE": "Material",
    "SWING STAGE - CONSUMABLES": "Material",
    "SWING STAGE - SAFETY EQUIPMENT": "Material",
    "SWING STAGE - INDIRECT COST": "Material",
    "CHANGE ORDER": "Other",
    "ENGINEERING / DRAWINGS": "Sub",
    "SWING STAGE - INSPECTIONS": "Labor",
    "PERMANENT SCAFFOLD SALE": "Material",
    "EQUIPMENT": "Material",
    "MOBILIZATION": "Labor",
    "DE-MOBILIZATION": "Labor",
    "SCAFFOLD RENTAL": "Rent",
    "CONSUMABLES / DIRECT MATERIAL": "Material",
    "LABOR - SCAFFOLD INSPECTIONS": "Labor",
    "SWITCH RAIL - LABOR": "Labor",
    "SWITCH RAIL - RENTAL": "Rent",
    "SWITCH RAIL - DELIVERY/PICK UP": "Delivery",
    "SITE TRUCK": "Delivery",
    "DELIVERY / PICK-UP": "Delivery",
    "TRASH CHUTE RENTAL": "Rent",
    "PER DIEM": "Labor",
    "TRAVEL": "Labor",
    "REPLACEMENT SCAFFOLD": "Material",
    "CONTRACT AMOUNT": "Other",
    "INDIRECT COST": "Other",
    "SUBCONTRACTOR CONTRACT AMOUNT": "Sub",
    "TEXTURA CHARGE": "Other",
    "CREDIT CARD PROCESSING": "Other",

    # Preserved from previous map just in case
    "SWING STAGE - MATERIAL": "Material",
    "SWING STAGE - DELIVERY": "Delivery",
    "SWING STAGE - SUB": "Sub",
    "SWING STAGE - OTHER": "Other",
    "SCAFFOLD - LABOR": "Labor",
    "SCAFFOLD - RENTAL": "Rent",
    "SCAFFOLD - MATERIAL": "Material",
    "SCAFFOLD - DELIVERY": "Delivery",
    "SCAFFOLD - SUB": "Sub",
    "SCAFFOLD - OTHER": "Other",
    "INSULATION - LABOR": "Labor",
    "INSULATION - MATERIAL": "Material",
    "INSULATION - DELIVERY": "Delivery",
    "INSULATION - SUB": "Sub",
    "INSULATION - OTHER": "Other",
    "RENTAL - SCAFFOLD": "Rent",
    "RENTAL - SWING STAGE": "Rent",
    "RENTAL - OTHER": "Other",
    "MATERIAL": "Material",
    "MATERIAL - SCAFFOLD": "Material",
    "MATERIAL - SWING STAGE": "Material",
    "MATERIAL - INSULATION": "Material",
    "DELIVERY": "Delivery",
    "DELIVERY - SCAFFOLD": "Delivery",
    "DELIVERY - SWING STAGE": "Delivery",
    "SUB": "Sub",
    "SUB - SCAFFOLD": "Sub",
    "SUB - SWING STAGE": "Sub",
    "OTHER": "Other",
}

ALLOWED_CATS = {"Labor", "Rent", "Material", "Sub", "Delivery", "Other"}

JOB_RE = re.compile(r"^[A-Z0-9]{4}-[A-Z0-9]{4,5}$", re.IGNORECASE)
INV_RE_5 = re.compile(r"\b(\d{5})\b")


def norm_text(x) -> str:
    if pd.isna(x):
        return ""
    return re.sub(r"\s+", " ", str(x).strip()).upper()


def is_job_code(x) -> bool:
    return bool(JOB_RE.match(str(x).strip()))


def is_swing_job(job_code: str) -> bool:
    if not isinstance(job_code, str) or "-" not in job_code:
        return False
    return job_code.split("-", 1)[1].upper().startswith("Z")


def find_header_row(path: Path, required: list[str], scan=300, sheet_name=0) -> int:
    preview = pd.read_excel(path, header=None, nrows=scan, sheet_name=sheet_name)
    req = [norm_text(r) for r in required]
    for i in range(len(preview)):
        row = [norm_text(v) for v in preview.iloc[i].tolist()]
        if all(any(r == cell for cell in row) for r in req):
            return i
    return 0


def build_invoice_to_job_map(gl_path: Path):
    gl_hdr = find_header_row(gl_path, required=["Code"], scan=300)
    gl = pd.read_excel(gl_path, header=gl_hdr)

    code_col = next((c for c in gl.columns if norm_text(c) == "CODE"), None)
    ref_col = next((c for c in gl.columns if norm_text(c) == "REFERENCE"), None)
    if code_col is None or ref_col is None:
        raise ValueError("GL: missing Code or Reference column")

    # per your rule: client is in column D
    client_col = gl.columns[3] if len(gl.columns) > 3 else gl.columns[0]

    df = gl[[code_col, ref_col, client_col]].copy()
    df.columns = ["Job Code", "Reference", "Client"]
    df["Job Code"] = df["Job Code"].astype(str).str.strip()
    df = df[df["Job Code"].apply(is_job_code)]

    df["Invoice#"] = df["Reference"].astype(str).str.extract(INV_RE_5, expand=False)
    df = df.dropna(subset=["Invoice#"])

    inv_to_job = df.drop_duplicates(subset=["Invoice#"])[["Invoice#", "Job Code"]]

    job_to_client = (
        df.groupby("Job Code", as_index=False)["Client"]
        .agg(lambda s: next((x for x in s if str(x).strip()), ""))
    )

    return inv_to_job, job_to_client


def parse_inventory_with_section_headers(inv_path: Path, inv_to_job: pd.DataFrame) -> pd.DataFrame:
    inv_hdr = find_header_row(inv_path, required=["Amount"], scan=350)
    inv = pd.read_excel(inv_path, header=inv_hdr)

    amount_col = next((c for c in inv.columns if norm_text(c) == "AMOUNT"), None)
    invoice_col = next((c for c in inv.columns if "INVOICE" in norm_text(c)), None)
    cust_id_col = next((c for c in inv.columns if norm_text(c) == "CUSTOMER ID"), None)
    cust_name_col = next((c for c in inv.columns if norm_text(c) == "CUSTOMER NAME"), None)

    if amount_col is None or invoice_col is None or cust_id_col is None:
        raise ValueError(f"Inventory: missing key columns. Columns: {list(inv.columns)}")

    if cust_name_col is None:
        # fallback if named column not found
        cust_name_col = inv.columns[3] if len(inv.columns) > 3 else inv.columns[0]

    # Keep minimum columns; first col is usually "Inv Date" or similar
    inv2 = inv[[inv.columns[0], invoice_col, cust_id_col, cust_name_col, amount_col]].copy()
    inv2.columns = ["Inv Date", "Invoice#", "Customer Id", "Customer Name", "Amount"]

    inv2["Invoice#"] = inv2["Invoice#"].astype(str).str.extract(INV_RE_5, expand=False)
    inv2["Amount"] = pd.to_numeric(inv2["Amount"], errors="coerce")

    valid_section_keys = {norm_text(k) for k in INTERNAL_CATEGORY_MAP.keys()}

    def is_blank(x):
        if pd.isna(x):
            return True
        return str(x).strip() == ""

    # Carry down last valid section header (only if it’s in our mapping)
    labels = []
    current_label = None

    for _, row in inv2.iterrows():
        cid_raw = "" if pd.isna(row["Customer Id"]) else str(row["Customer Id"]).strip()
        cid = norm_text(cid_raw)

        is_header = (
            pd.isna(row["Amount"])
            and is_blank(row["Invoice#"])
            and is_blank(row["Customer Name"])
            and cid in valid_section_keys
        )

        if is_header:
            current_label = cid_raw

        labels.append(current_label)

    inv2["Section"] = labels

    # Keep only transaction lines: Amount present and Invoice present
    tx = inv2.dropna(subset=["Amount"]).copy()
    tx = tx[~tx["Invoice#"].isna()].copy()

    tx["Client"] = tx["Customer Name"].astype(str).str.strip().replace("nan", "")
    tx["Section"] = tx["Section"].fillna("").apply(norm_text)

    # map section -> category
    normalized_map = {norm_text(k): v for k, v in INTERNAL_CATEGORY_MAP.items()}
    tx["Category"] = tx["Section"].map(normalized_map).fillna("Other")

    # map invoice -> job
    tx = tx.merge(inv_to_job, on="Invoice#", how="left").dropna(subset=["Job Code"])
    tx["Job Code"] = tx["Job Code"].astype(str).str.strip()

    return tx


def build_revenue_summary(tx: pd.DataFrame, job_to_client: pd.DataFrame) -> pd.DataFrame:
    rev = (
        tx.groupby(["Job Code", "Client", "Category"], as_index=False)["Amount"].sum()
        .pivot_table(
            index=["Job Code", "Client"],
            columns="Category",
            values="Amount",
            aggfunc="sum"
        )
        .fillna(0.0)
        .reset_index()
    )

    for col in ["Labor", "Rent", "Material", "Sub", "Delivery", "Other"]:
        if col not in rev.columns:
            rev[col] = 0.0

    rev["Total"] = rev[["Labor", "Rent", "Material", "Sub", "Delivery", "Other"]].sum(axis=1)
    rev = rev[["Job Code", "Client", "Labor", "Rent", "Material", "Sub", "Delivery", "Other", "Total"]]

    # fallback client from GL if blank
    rev = rev.merge(job_to_client, on="Job Code", how="left", suffixes=("", "_GL"))
    rev["Client"] = rev["Client"].where(rev["Client"].astype(str).str.strip() != "", rev["Client_GL"].fillna(""))
    rev = rev.drop(columns=["Client_GL"])

    return rev


def build_expense_summary(cost_path: Path) -> pd.DataFrame:
    cost_hdr = find_header_row(cost_path, required=["Amount"], scan=300)
    jc = pd.read_excel(cost_path, header=cost_hdr)

    cols = list(jc.columns)
    jobc = cols[0]
    codec = cols[1] if len(cols) > 1 else cols[0]

    # per your rule: amount is column F, but still try to find "Amount" header first
    amtc = next((c for c in cols if norm_text(c) == "AMOUNT"), None)
    if amtc is None:
        amtc = cols[5] if len(cols) > 5 else cols[-1]

    df = jc[[jobc, codec, amtc]].copy()
    df.columns = ["Job Code", "Cost Code", "Amount"]
    df["Job Code"] = df["Job Code"].astype(str).str.strip()
    df = df[df["Job Code"].apply(is_job_code)]

    df["Cost Code"] = df["Cost Code"].astype(str).str.strip().str.upper()
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0.0)

    df["Bucket"] = df["Cost Code"].apply(lambda x: "Labor" if x in {"L", "PD", "T"} else "Other Direct Costs")

    exp = (
        df.groupby(["Job Code", "Bucket"], as_index=False)["Amount"].sum()
        .pivot_table(index="Job Code", columns="Bucket", values="Amount", aggfunc="sum")
        .fillna(0.0)
        .reset_index()
    )

    for col in ["Labor", "Other Direct Costs"]:
        if col not in exp.columns:
            exp[col] = 0.0

    exp["Total"] = exp[["Labor", "Other Direct Costs"]].sum(axis=1)
    return exp[["Job Code", "Labor", "Other Direct Costs", "Total"]]


def load_job_master(jobmaster_path: Path) -> pd.DataFrame:
    raw = pd.read_excel(jobmaster_path, header=None, nrows=300)
    hdr = 0
    for i in range(len(raw)):
        if norm_text(raw.iloc[i, 0]) == "JOB":
            hdr = i
            break

    jm = pd.read_excel(jobmaster_path, header=hdr)
    cols = list(jm.columns)

    out = jm[[cols[0], cols[1], cols[2]]].copy()
    out.columns = ["Job Code", "Description", "Sales Rep"]
    out["Job Code"] = out["Job Code"].astype(str).str.strip()
    out = out[out["Job Code"].apply(is_job_code)].drop_duplicates("Job Code")

    out["Description"] = out["Description"].astype(str).str.strip().replace("nan", "")
    out["Sales Rep"] = out["Sales Rep"].astype(str).str.strip().replace("nan", "")
    return out


def build_labor_analysis(rev: pd.DataFrame, exp: pd.DataFrame, jm: pd.DataFrame) -> pd.DataFrame:
    df = rev.merge(exp, on="Job Code", how="left", suffixes=("", "_COST")).merge(jm, on="Job Code", how="left")

    df["Labor Revenue"] = df["Labor"].fillna(0.0)
    df["Labor Cost"] = df["Labor_COST"].fillna(0.0) if "Labor_COST" in df.columns else 0.0

    # remove jobs where both are zero
    df = df[~((df["Labor Revenue"] == 0) & (df["Labor Cost"] == 0))].copy()

    # swing-stage exclusion
    df["Is Swing"] = df["Job Code"].apply(is_swing_job)
    df = df[~((df["Is Swing"]) & (df["Labor Revenue"] > 0) & (df["Labor Cost"] == 0))].copy()

    # billing check
    df["Billing Check"] = ""
    df.loc[(df["Labor Cost"] > 0) & (df["Labor Revenue"] == 0), "Billing Check"] = "NO BILLING"
    df.loc[(df["Labor Cost"] == 0) & (df["Labor Revenue"] > 0), "Billing Check"] = "No Costs"

    def margin(rev_amt, cost_amt):
        if rev_amt == 0 and cost_amt > 0:
            return 0.0
        if cost_amt == 0 and rev_amt > 0:
            return 1.0
        if rev_amt == 0:
            return None
        return (rev_amt - cost_amt) / rev_amt

    df["Labor Margin %"] = [margin(r, c) for r, c in zip(df["Labor Revenue"], df["Labor Cost"])]

    out = df[
        ["Job Code", "Description", "Client", "Labor Revenue", "Labor Cost", "Labor Margin %", "Billing Check", "Sales Rep"]
    ].copy()

    out[["Description", "Client", "Sales Rep"]] = out[["Description", "Client", "Sales Rep"]].fillna("")
    return out


def format_workbook(path: Path):
    wb = load_workbook(path)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(wrap_text=True, vertical="center")

    dollar_fmt = "$#,##0"
    pct_fmt = "0.0%"

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Revenue formats
    ws = wb["Revenue Summary"]
    cm = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for name in ["Labor", "Rent", "Material", "Sub", "Delivery", "Other", "Total"]:
        c = cm.get(name)
        if c:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = dollar_fmt

    # Expense formats
    ws = wb["Expense Summary"]
    cm = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for name in ["Labor", "Other Direct Costs", "Total"]:
        c = cm.get(name)
        if c:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = dollar_fmt

    # Labor Analysis formats + conditional formatting
    ws = wb["Labor Analysis"]
    cm = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    for name in ["Labor Revenue", "Labor Cost"]:
        c = cm.get(name)
        if c:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = dollar_fmt

    c = cm.get("Labor Margin %")
    if c:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = pct_fmt

    def setw(col_name, w):
        if col_name in cm:
            ws.column_dimensions[get_column_letter(cm[col_name])].width = w

    setw("Job Code", 14)
    setw("Description", 44)
    setw("Client", 28)
    setw("Billing Check", 14)
    setw("Sales Rep", 18)
    setw("Labor Revenue", 16)
    setw("Labor Cost", 14)
    setw("Labor Margin %", 14)

    if ws.max_row >= 2:
        ws.conditional_formatting._cf_rules.clear()

        margin_letter = get_column_letter(cm["Labor Margin %"])
        billing_letter = get_column_letter(cm["Billing Check"])
        last_col_letter = get_column_letter(ws.max_column)

        data_start = 2
        data_range = f"A{data_start}:{last_col_letter}{ws.max_row}"

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'OR(${billing_letter}{data_start}="NO BILLING",${margin_letter}{data_start}<0)'],
                fill=red_fill,
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'AND(${margin_letter}{data_start}>=0,${margin_letter}{data_start}<=0.15)'],
                fill=orange_fill,
            ),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'AND(${margin_letter}{data_start}>0.15,${margin_letter}{data_start}<=0.25)'],
                fill=yellow_fill,
            ),
        )

    wb.save(path)


def build_report(gl_path: str, inventory_path: str, job_cost_path: str, job_master_path: str, output_path: str) -> str:
    # We rename your run_report function to build_report to match the GUI exactly
    gl_path = Path(gl_path)
    inv_path = Path(inventory_path)
    cost_path = Path(job_cost_path)
    jobmaster_path = Path(job_master_path)
    output_path = Path(output_path)

    # basic sanity on mapping
    for k, v in INTERNAL_CATEGORY_MAP.items():
        if v not in ALLOWED_CATS:
            raise ValueError(f"Bad category mapping value: {k} -> {v} (must be one of {sorted(ALLOWED_CATS)})")

    inv_to_job, job_to_client = build_invoice_to_job_map(gl_path)
    tx = parse_inventory_with_section_headers(inv_path, inv_to_job)

    revenue_summary = build_revenue_summary(tx, job_to_client)
    expense_summary = build_expense_summary(cost_path)
    job_master = load_job_master(jobmaster_path)

    labor_analysis = build_labor_analysis(revenue_summary, expense_summary, job_master)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        revenue_summary.to_excel(writer, sheet_name="Revenue Summary", index=False)
        expense_summary.to_excel(writer, sheet_name="Expense Summary", index=False)
        labor_analysis.to_excel(writer, sheet_name="Labor Analysis", index=False)

    format_workbook(output_path)
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate CD job revenue/cost + labor margin analysis workbook.")
    parser.add_argument("--gl", required=True, help="Path to General Ledger.xlsx")
    parser.add_argument("--inv", required=True, help="Path to Sales Analysis by Inventory Item.xlsx")
    parser.add_argument("--cost", required=True, help="Path to Job Cost Query.xlsx")
    parser.add_argument("--jobmaster", required=True, help="Path to Job Summary of Billings and Cost.xlsx")
    parser.add_argument("--out", default="job_labor_analysis_output.xlsx", help="Output xlsx path")

    args = parser.parse_args()

    build_report(
        gl_path=args.gl,
        inventory_path=args.inv,
        job_cost_path=args.cost,
        job_master_path=args.jobmaster,
        output_path=args.out,
    )
    print(f"Done -> {args.out}")


if __name__ == "__main__":
    main()
